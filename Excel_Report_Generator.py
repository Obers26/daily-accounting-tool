import pandas as pd
import sqlite3
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Alignment
import calendar
from typing import Tuple, Dict, List, Optional

# Constants
LIGHT_BLUE_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="B3E19A", end_color="B3E19A", fill_type="solid")
DATE_FORMAT = '%m/%d/%Y'
PERCENTAGE_FORMAT = '0.00%'
NUMBER_FORMAT = '#,##0;(#,##0)'

# Column mappings for Overall sheet
OVERALL_COLUMNS = {
    'TOTAL_PL': 'F',
    'PERIOD_STARTING_NAV': 'G',
    'START_FUND_VALUE_ACCOUNTS': 'H',
    'END_FUND_VALUE_ACCOUNTS': 'I',
    'START_FUND_VALUE_NAV_CUM_PL': 'J',
    'END_FUND_VALUE_NAV_CUM_PL': 'K',
    'DAILY_FUND_RETURN': 'L',
    'PERIOD_CUMULATIVE_PL': 'M',
    'PERIOD_CUMULATIVE_RETURN': 'N'
}

# Column mappings for Broker sheet
BROKER_COLUMNS = {
    'PL': 'B',
    'INTEREST': 'G',
    'DIVIDENDS': 'H',
    'DEPOSITS_WITHDRAWALS': 'I',
    'TOTAL_BROKER': 'L'
}


class ExcelReportGenerator:
    """Generate Excel reports from daily accounting database."""
    
    def __init__(self, db_path: str = 'daily_accounting.db'):
        self.db_path = db_path
        self.conn = None
    
    def _connect_to_database(self) -> bool:
        """Connect to the SQLite database."""
        try:
            self.conn = sqlite3.connect(self.db_path)
            return True
        except sqlite3.Error as e:
            print(f"Database connection error: {e}")
            return False
    
    def _close_database(self):
        """Close database connection."""
        if self.conn:
            self.conn.close()
    
    def _validate_dates(self, start_date: str, end_date: str) -> Tuple[bool, str, Optional[datetime], Optional[datetime]]:
        """Validate input dates."""
        try:
            start_dt = datetime.strptime(start_date, DATE_FORMAT)
            end_dt = datetime.strptime(end_date, DATE_FORMAT)
            
            if start_dt > end_dt:
                return False, "Start date must be before end date", None, None
            
            return True, "", start_dt, end_dt
        except ValueError:
            return False, "Invalid date format. Use MM/DD/YYYY", None, None
    
    def _query_data(self, table_name: str, start_date: str, end_date: str) -> pd.DataFrame:
        """Query data from specified table for date range."""
        query = f'''
            SELECT * FROM {table_name} 
            WHERE "Date" BETWEEN ? AND ?
            ORDER BY "Date"
        '''
        return pd.read_sql_query(query, self.conn, params=(start_date, end_date))
    
    def _prepare_dataframes(self, start_date: str, end_date: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """Prepare all required dataframes."""
        # Query broker data
        df_broker = self._query_data('broker', start_date, end_date)
        if df_broker.empty:
            raise ValueError("No broker data found for the specified date range")
        
        # Query overall data
        df_overall = self._query_data('overall', start_date, end_date)
        
        # Query other transactions data
        df_other = self._query_data('other_transactions', start_date, end_date)
        
        # Clean and process dataframes
        df_other = self._clean_other_transactions(df_other)
        df_broker, df_overall, df_other = self._set_date_indices(df_broker, df_overall, df_other)
        
        return df_broker, df_overall, df_other
    
    def _clean_other_transactions(self, df_other: pd.DataFrame) -> pd.DataFrame:
        """Clean and format other transactions data."""
        if df_other.empty:
            return df_other
        
        # Remove unwanted columns
        if 'id' in df_other.columns:
            df_other = df_other.drop('id', axis=1)
        
        # Convert binary columns to Yes/No
        binary_columns = ['Counted in P&L', 'Overnight']
        for col in binary_columns:
            if col in df_other.columns:
                df_other[col] = df_other[col].map({1: 'Yes', 0: 'No'})
        
        return df_other
    
    def _set_date_indices(self, df_broker: pd.DataFrame, df_overall: pd.DataFrame, 
                         df_other: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """Set date as index for all dataframes and add calculated columns."""
        df_broker.set_index('Date', inplace=True)
        
        if not df_overall.empty:
            df_overall.set_index('Date', inplace=True)
            # Add calculated columns
            df_overall['Daily Fund Return'] = 0.0
            df_overall['Period Cumulative P&L'] = 0.0
            df_overall['Period Cumulative Return'] = 0.0
        
        if not df_other.empty:
            df_other.set_index('Date', inplace=True)
        
        return df_broker, df_overall, df_other
    
    def _create_overall_sheet(self, writer: pd.ExcelWriter, df_overall: pd.DataFrame):
        """Create and format the Overall sheet."""
        if df_overall.empty:
            return
        
        df_overall.to_excel(writer, sheet_name='Overall')
        worksheet = writer.sheets['Overall']
        
        self._format_overall_sheet(worksheet, df_overall)
        self._add_overall_formulas(worksheet, df_overall)
        self._highlight_special_rows(worksheet, df_overall)
    
    def _format_overall_sheet(self, worksheet, df_overall: pd.DataFrame):
        """Apply basic formatting to Overall sheet."""
        # Freeze panes
        worksheet.freeze_panes = 'B2'
        
        # Apply multi-line header formatting
        self._apply_header_formatting(worksheet, df_overall, start_row=1)
        
        # Set date column width
        worksheet.column_dimensions['A'].width = 12
        
        # Auto-adjust column widths and format numbers
        for idx, col in enumerate(df_overall.columns):
            # Use the new column width calculation method
            column_width = self._calculate_column_width(df_overall[col], str(col))
            worksheet.column_dimensions[chr(65 + idx + 1)].width = column_width
            
            # Format numbers in the column
            for row in range(2, len(df_overall) + 2):
                cell = worksheet.cell(row=row, column=idx + 2)
                if isinstance(cell.value, (int, float)):
                    if col in ['Daily Fund Return', 'Period Cumulative Return']:
                        cell.number_format = PERCENTAGE_FORMAT
                    else:
                        cell.number_format = NUMBER_FORMAT
    
    def _add_overall_formulas(self, worksheet, df_overall: pd.DataFrame):
        """Add Excel formulas to Overall sheet."""
        cols = OVERALL_COLUMNS
        
        for row in range(2, len(df_overall) + 2):
            # Daily Fund Return formula
            daily_return_cell = f'{cols["DAILY_FUND_RETURN"]}{row}'
            numerator = f'{cols["TOTAL_PL"]}{row}'
            denominator = f'{cols["START_FUND_VALUE_NAV_CUM_PL"]}{row}'
            worksheet[daily_return_cell] = f'=({numerator}/{denominator})'
            worksheet[daily_return_cell].number_format = PERCENTAGE_FORMAT
            
            # Period Cumulative P&L formula
            cumulative_pl_cell = f'{cols["PERIOD_CUMULATIVE_PL"]}{row}'
            if row == 2:
                worksheet[cumulative_pl_cell] = f'={cols["TOTAL_PL"]}{row}'
            else:
                worksheet[cumulative_pl_cell] = (
                    f'=SUMIFS({cols["TOTAL_PL"]}$2:{cols["TOTAL_PL"]}{row},'
                    f'{cols["PERIOD_STARTING_NAV"]}$2:{cols["PERIOD_STARTING_NAV"]}{row},'
                    f'{cols["PERIOD_STARTING_NAV"]}{row})'
                )
            worksheet[cumulative_pl_cell].number_format = NUMBER_FORMAT
            
            # Period Cumulative Return formula
            cumulative_return_cell = f'{cols["PERIOD_CUMULATIVE_RETURN"]}{row}'
            worksheet[cumulative_return_cell] = (
                f'=({cols["PERIOD_CUMULATIVE_PL"]}{row}/{cols["PERIOD_STARTING_NAV"]}{row})'
            )
            worksheet[cumulative_return_cell].number_format = PERCENTAGE_FORMAT
    
    def _highlight_special_rows(self, worksheet, df_overall: pd.DataFrame):
        """Highlight month-end and valuation day rows."""
        for row_idx, (date_str, row_data) in enumerate(df_overall.iterrows()):
            excel_row = row_idx + 2
            
            try:
                date_obj = datetime.strptime(date_str, DATE_FORMAT)
            except ValueError:
                continue
            
            is_month_end = self._is_month_end(row_idx, date_obj, df_overall)
            is_valuation_day = self._is_valuation_day(row_idx, row_data, df_overall)
            
            # Apply highlighting (valuation day takes precedence)
            fill_color = None
            if is_valuation_day:
                fill_color = LIGHT_GREEN_FILL
            elif is_month_end:
                fill_color = LIGHT_BLUE_FILL
            
            if fill_color:
                for col_idx in range(len(df_overall.columns) + 1):
                    cell = worksheet.cell(row=excel_row, column=col_idx + 1)
                    cell.fill = fill_color
    
    def _is_month_end(self, row_idx: int, date_obj: datetime, df_overall: pd.DataFrame) -> bool:
        """Check if date is month-end."""
        if row_idx == len(df_overall) - 1:
            return True
        
        try:
            next_date_str = df_overall.index[row_idx + 1]
            next_date_obj = datetime.strptime(next_date_str, DATE_FORMAT)
            return (next_date_obj.month != date_obj.month or 
                   next_date_obj.year != date_obj.year)
        except (ValueError, IndexError):
            return False
    
    def _is_valuation_day(self, row_idx: int, row_data: pd.Series, df_overall: pd.DataFrame) -> bool:
        """Check if date is a valuation day."""
        if row_idx == 0:
            return True
        
        current_nav = row_data['Period Starting NAV']
        prev_nav = df_overall.iloc[row_idx - 1]['Period Starting NAV']
        return current_nav != prev_nav
    
    def _create_period_returns_sheet(self, writer: pd.ExcelWriter, df_overall: pd.DataFrame):
        """Create Period Returns sheet."""
        if df_overall.empty:
            return
        
        period_returns_data = self._extract_period_returns_data(df_overall)
        if not period_returns_data:
            return
        
        df_period_returns = pd.DataFrame(period_returns_data)
        df_period_returns.to_excel(writer, sheet_name='Period Returns', index=False)
        
        worksheet = writer.sheets['Period Returns']
        self._format_period_returns_sheet(worksheet, df_period_returns)
    
    def _extract_period_returns_data(self, df_overall: pd.DataFrame) -> List[Dict]:
        """Extract period returns data from overall dataframe."""
        period_returns_data = []
        current_period_start_nav = None
        
        for row_idx, (date_str, row_data) in enumerate(df_overall.iterrows()):
            current_nav = row_data['Period Starting NAV']
            
            if current_period_start_nav is None or current_nav != current_period_start_nav:
                current_period_start_nav = current_nav
            
            # Check if this is the end of the current period
            is_period_end = (row_idx == len(df_overall) - 1 or 
                           df_overall.iloc[row_idx + 1]['Period Starting NAV'] != current_nav)
            
            if is_period_end:
                period_returns_data.append({
                    'Period End Date': date_str,
                    'Period Starting NAV': current_period_start_nav,
                    'P&L': 0,
                    'Fund Return': 0
                })
        
        return period_returns_data
    
    def _format_period_returns_sheet(self, worksheet, df_period_returns: pd.DataFrame):
        """Format the Period Returns sheet."""
        worksheet.freeze_panes = 'A2'
        
        # Apply multi-line header formatting (no index column for this sheet)
        self._apply_header_formatting(worksheet, df_period_returns, start_row=1, has_index=False)
        
        # Add Excel formulas
        for row in range(2, len(df_period_returns) + 2):
            period_end_date = df_period_returns.iloc[row - 2]['Period End Date']
            
            # VLOOKUP formulas for NAV, P&L, and Return
            worksheet[f'B{row}'] = f'=VLOOKUP("{period_end_date}",Overall!A:G,7,FALSE)'
            worksheet[f'B{row}'].number_format = NUMBER_FORMAT
            
            worksheet[f'C{row}'] = f'=VLOOKUP("{period_end_date}",Overall!A:M,13,FALSE)'
            worksheet[f'C{row}'].number_format = NUMBER_FORMAT
            
            worksheet[f'D{row}'] = f'=VLOOKUP("{period_end_date}",Overall!A:N,14,FALSE)'
            worksheet[f'D{row}'].number_format = PERCENTAGE_FORMAT
        
        # Auto-adjust column widths using the new method
        for idx, col in enumerate(df_period_returns.columns):
            if idx == 0:  # First column (dates)
                worksheet.column_dimensions[chr(65 + idx)].width = 15
            else:
                # Create a dummy series for width calculation
                dummy_series = pd.Series([0] * len(df_period_returns))
                column_width = self._calculate_column_width(dummy_series, str(col))
                worksheet.column_dimensions[chr(65 + idx)].width = max(column_width, 12)
    
    def _create_broker_sheet(self, writer: pd.ExcelWriter, df_broker: pd.DataFrame):
        """Create and format the Brokerage Account sheet."""
        df_broker.to_excel(writer, sheet_name='Brokerage Account')
        worksheet = writer.sheets['Brokerage Account']
        
        self._format_broker_sheet(worksheet, df_broker)
        self._add_broker_formulas(worksheet, df_broker)
        self._validate_broker_calculations(df_broker)
    
    def _format_broker_sheet(self, worksheet, df_broker: pd.DataFrame):
        """Apply basic formatting to Brokerage Account sheet."""
        worksheet.freeze_panes = 'A2'
        
        # Apply multi-line header formatting
        self._apply_header_formatting(worksheet, df_broker, start_row=1)
        
        # Set date column width
        worksheet.column_dimensions['A'].width = 12
        
        # Auto-adjust column widths and format numbers
        for idx, col in enumerate(df_broker.columns):
            # Use the new column width calculation method
            column_width = self._calculate_column_width(df_broker[col], str(col))
            worksheet.column_dimensions[chr(65 + idx + 1)].width = column_width
            
            # Format numbers
            for row in range(2, len(df_broker) + 2):
                cell = worksheet.cell(row=row, column=idx + 2)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = NUMBER_FORMAT
    
    def _add_broker_formulas(self, worksheet, df_broker: pd.DataFrame):
        """Add P&L calculation formulas to Brokerage Account sheet."""
        cols = BROKER_COLUMNS
        
        for row in range(3, len(df_broker) + 2):
            # P&L = Total Broker - Previous Total Broker - Deposits - Dividends - Interest
            pnl_cell = f'{cols["PL"]}{row}'
            formula = (
                f'={cols["TOTAL_BROKER"]}{row}-{cols["TOTAL_BROKER"]}{row-1}-'
                f'{cols["DEPOSITS_WITHDRAWALS"]}{row}-{cols["DIVIDENDS"]}{row}-'
                f'{cols["INTEREST"]}{row}'
            )
            worksheet[pnl_cell] = formula
            worksheet[pnl_cell].number_format = NUMBER_FORMAT
    
    def _validate_broker_calculations(self, df_broker: pd.DataFrame):
        """Validate P&L calculations against database values."""
        print("Validating P&L calculations...")
        tolerance = 0.01
        discrepancies_found = False
        original_pnl_values = df_broker['P&L'].copy()
        
        for idx, (date_idx, row_data) in enumerate(df_broker.iterrows()):
            if idx == 0:
                continue
            
            # Calculate P&L manually
            current_total = row_data['Total Broker'] or 0
            prev_total = df_broker.iloc[idx-1]['Total Broker'] or 0
            deposits_withdrawals = row_data['Deposits & Withdrawals'] or 0
            dividends = row_data['Dividends'] or 0
            interest = row_data['Interest'] or 0
            
            calculated_value = current_total - prev_total - deposits_withdrawals - dividends - interest
            database_value = original_pnl_values.iloc[idx]
            
            if database_value is not None:
                if abs(calculated_value - database_value) > tolerance:
                    discrepancies_found = True
                    self._print_discrepancy(date_idx, database_value, calculated_value, 
                                          current_total, prev_total, deposits_withdrawals, 
                                          dividends, interest)
            else:
                print(f"P&L DISCREPANCY for {date_idx}: Database P&L is None")
        
        if not discrepancies_found:
            print("✓ All P&L calculations match between database and Excel formulas")
        else:
            print("⚠ P&L discrepancies detected - please review the data")
    
    def _print_discrepancy(self, date_idx: str, database_value: float, calculated_value: float,
                          current_total: float, prev_total: float, deposits_withdrawals: float,
                          dividends: float, interest: float):
        """Print P&L discrepancy details."""
        print(f"P&L DISCREPANCY for {date_idx}:")
        print(f"  Database P&L: ${database_value:.2f}")
        print(f"  Formula P&L: ${calculated_value:.2f}")
        print(f"  Difference: ${abs(calculated_value - database_value):.2f}")
        print(f"  Components: Total=${current_total:.2f}, PrevTotal=${prev_total:.2f}, "
              f"Deposits=${deposits_withdrawals:.2f}, Dividends=${dividends:.2f}, "
              f"Interest=${interest:.2f}")
    
    def _create_other_transactions_sheet(self, writer: pd.ExcelWriter, df_other: pd.DataFrame):
        """Create and format the Other Transactions sheet."""
        if df_other.empty:
            return
        
        df_other.to_excel(writer, sheet_name='Other Transactions')
        worksheet = writer.sheets['Other Transactions']
        
        # Format sheet
        worksheet.freeze_panes = 'A2'
        
        # Apply multi-line header formatting
        self._apply_header_formatting(worksheet, df_other, start_row=1)
        
        # Set date column width
        worksheet.column_dimensions['A'].width = 12
        
        # Auto-adjust column widths and format numbers
        for idx, col in enumerate(df_other.columns):
            # Use the new column width calculation method
            column_width = self._calculate_column_width(df_other[col], str(col))
            worksheet.column_dimensions[chr(65 + idx + 1)].width = column_width
            
            # Format numbers
            for row in range(2, len(df_other) + 2):
                cell = worksheet.cell(row=row, column=idx + 2)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = NUMBER_FORMAT
    
    def _split_header_text(self, header: str, max_length: int = 15) -> str:
        """Split long header text into two lines for better column width."""
        if len(header) <= max_length:
            return header
        
        # Try to split at natural break points
        words = header.split()
        if len(words) <= 1:
            return header
        
        # Find the best split point
        mid_point = len(words) // 2
        first_line = ' '.join(words[:mid_point])
        second_line = ' '.join(words[mid_point:])
        
        # If either line is still too long, try a different split
        if len(first_line) > max_length or len(second_line) > max_length:
            # Try splitting at common separators
            for separator in [' & ', ' - ', ' / ', ' ']:
                if separator in header:
                    parts = header.split(separator, 1)
                    if len(parts) == 2 and len(parts[0]) <= max_length and len(parts[1]) <= max_length:
                        return f"{parts[0]}\n{parts[1]}"
        
        return f"{first_line}\n{second_line}"
    
    def _apply_header_formatting(self, worksheet, df: pd.DataFrame, start_row: int = 1, has_index: bool = True):
        """Apply multi-line header formatting to worksheet."""
        # Set header row height to accommodate two lines
        worksheet.row_dimensions[start_row].height = 30
        
        # Determine column offset based on whether there's an index column
        col_offset = 2 if has_index else 1
        
        # Format header cells
        for col_idx, column_name in enumerate(df.columns):
            cell = worksheet.cell(row=start_row, column=col_idx + col_offset)
            cell.value = self._split_header_text(str(column_name))
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        
        # Also format the index column header if it exists
        if has_index and hasattr(df.index, 'name') and df.index.name:
            cell = worksheet.cell(row=start_row, column=1)
            cell.value = self._split_header_text(str(df.index.name))
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    def _calculate_column_width(self, column_data: pd.Series, header: str) -> int:
        """Calculate optimal column width considering wrapped headers."""
        # Get the maximum length of the wrapped header lines
        wrapped_header = self._split_header_text(str(header))
        header_lines = wrapped_header.split('\n')
        max_header_length = max(len(line) for line in header_lines)
        
        # Get the maximum length of the data
        max_data_length = column_data.astype(str).apply(len).max()
        
        # Return the maximum of header and data, with some padding
        return max(max_header_length, max_data_length) + 2
    
    def generate_excel_report(self, start_date: str, end_date: str, output_path: str) -> Tuple[bool, str]:
        """
        Generate an Excel report for the specified date range.
        
        Args:
            start_date: Start date in format 'MM/DD/YYYY'
            end_date: End date in format 'MM/DD/YYYY'
            output_path: Path where the Excel file should be saved
        
        Returns:
            Tuple of (success: bool, message: str)
        """
        try:
            # Validate inputs
            valid, error_msg, start_dt, end_dt = self._validate_dates(start_date, end_date)
            if not valid:
                return False, error_msg
            
            # Connect to database
            if not self._connect_to_database():
                return False, "Failed to connect to database"
            
            # Prepare data
            df_broker, df_overall, df_other = self._prepare_dataframes(start_date, end_date)
            
            # Create Excel file
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                self._create_overall_sheet(writer, df_overall)
                self._create_period_returns_sheet(writer, df_overall)
                self._create_broker_sheet(writer, df_broker)
                self._create_other_transactions_sheet(writer, df_other)
            
            # Generate success message
            sheets_created = self._get_sheets_created(df_overall, df_other)
            success_message = f"Excel report generated successfully with sheets: {', '.join(sheets_created)}"
            
            return True, success_message
            
        except Exception as e:
            return False, f"Error generating report: {str(e)}"
        finally:
            self._close_database()
    
    def _get_sheets_created(self, df_overall: pd.DataFrame, df_other: pd.DataFrame) -> List[str]:
        """Get list of sheets that were created."""
        sheets_created = []
        if not df_overall.empty:
            sheets_created.extend(['Overall', 'Period Returns'])
        sheets_created.append('Brokerage Account')
        if not df_other.empty:
            sheets_created.append('Other Transactions')
        return sheets_created